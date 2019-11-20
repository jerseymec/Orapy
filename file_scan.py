import openpyxl
import os
target_location = os.getcwd()

print(target_location)
data_wb = openpyxl.load_workbook('C:\\Users\\pereira_charles\\Onbase\\lookup\\Falcon7x.xlsx',read_only=True)

#worksheet = data_wb.get_sheet_by_name('Sheet1')
worksheet = data_wb['Sheet1']

dict = {}


for rownum in range(2,worksheet.max_row ):
    #print(worksheet.cell(row=rownum, column = 2).value)
    line  = worksheet.cell(row=rownum, column = 2).value
   # print(line)
    pdf_names = line.split()[-1]
    dict.setdefault(pdf_names,line)

print(worksheet.max_row)
row=0
for k,v in dict.items():
    row+=1
  #  print(" {} : {} ".format(k,v))
print(row)